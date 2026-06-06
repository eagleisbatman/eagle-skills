#!/usr/bin/env python3
"""Build a project tracker + Gantt workbook (.xlsx) that opens in Excel AND
imports cleanly into Google Sheets.

Input is a JSON spec (see assets/example-spec.json). Output is a styled workbook
with three sheets:
  Tracker   — task table: status dropdown (data validation), computed Duration,
              %-complete data bars, status-colored rows.
  Gantt     — conditional-formatting cell-grid Gantt; bars colored by status,
              today marker on the header row. No macros, no native chart object,
              so it round-trips to Google Sheets.
  Dashboard — three rollups: total tasks, weighted % complete, overdue count.

Dates are coerced to real dates and written with a date number_format so the
Gantt CF formulas (which compare cells to TODAY()/EOMONTH) actually fire — a
date stored as text silently breaks the bars, so we fail loudly instead.

Usage:
  python build_tracker.py spec.json [-o tracker.xlsx]
                                    [--timescale day|week|month]
                                    [--sheets-native]   # adds a SPARKLINE column
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---- Clean-sheet design tokens (Helvetica, monochrome) -----------------------
FB = Font(name="Helvetica", size=10.5, color="000000")
FH = Font(name="Helvetica", size=10, color="000000", bold=True)
FT = Font(name="Helvetica", size=14, color="000000", bold=True)
FCAP = Font(name="Helvetica", size=8.5, color="6B7280")
FILL_H = PatternFill("solid", fgColor="F3F4F6")
FILL_W = PatternFill("solid", fgColor="FFFFFF")
TH = Side(style="thin", color="D1D5DB")
BD = Border(top=TH, bottom=TH, left=TH, right=TH)
AL = Alignment(horizontal="left", vertical="top", wrap_text=True)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AH = Alignment(horizontal="left", vertical="center", wrap_text=True)
DATE_FMT = "yyyy-mm-dd"

# Functional status palette — color carries MEANING (RAG), so it is kept even
# though the base clean-sheet system is monochrome. (light fill, text, bar).
STATUS = {
    "Not started": ("E5E7EB", "374151", "D1D5DB"),
    "In progress": ("FEF3C7", "92400E", "F59E0B"),
    "Blocked":     ("FEE2E2", "991B1B", "EF4444"),
    "At risk":     ("FFEDD5", "9A3412", "FB923C"),
    "On hold":     ("E2E8F0", "475569", "94A3B8"),
    "Done":        ("D1FAE5", "065F46", "10B981"),
    "Milestone":   ("E5E7EB", "111827", "1F2937"),
}
DEFAULT_STATUSES = ["Not started", "In progress", "Blocked", "On hold", "Done"]


def _date(value: str, field: str, task: str) -> dt.date:
    """Coerce 'YYYY-MM-DD' to a date. Fail loudly — a text date breaks the Gantt."""
    if isinstance(value, dt.date):
        return value
    try:
        return dt.date.fromisoformat(str(value))
    except (TypeError, ValueError):
        raise SystemExit(
            f"Bad {field} date {value!r} on task {task!r} — use YYYY-MM-DD."
        ) from None


def _buckets(lo: dt.date, hi: dt.date, scale: str) -> list[dt.date]:
    """Period-start dates spanning [lo, hi] at day/week/month granularity."""
    if scale == "day":
        n = (hi - lo).days
        return [lo + dt.timedelta(days=i) for i in range(n + 1)]
    if scale == "week":
        start = lo - dt.timedelta(days=lo.weekday())  # back to Monday
        out, cur = [], start
        while cur <= hi:
            out.append(cur)
            cur += dt.timedelta(days=7)
        return out
    out, cur = [], dt.date(lo.year, lo.month, 1)
    while cur <= hi:
        out.append(cur)
        cur = dt.date(cur.year + (cur.month == 12), (cur.month % 12) + 1, 1)
    return out


def _bucket_end(col: str, scale: str) -> str:
    """CF expression for the inclusive end-date of the bucket whose start date is
    in header cell `col`3 (the bucket header lives on the absolute row 3)."""
    return {"day": f"{col}$3", "week": f"{col}$3+6"}.get(
        scale, f"EOMONTH({col}$3,0)")


def _norm(tasks: list[dict]) -> list[dict]:
    for t in tasks:
        t["_start"] = _date(t["start"], "start", t.get("task", "?"))
        t["_end"] = _date(t["end"], "end", t.get("task", "?"))
        if t["_end"] < t["_start"]:
            raise SystemExit(f"end before start on task {t.get('task')!r}")
        t.setdefault("percent", 0)
    return tasks


def _hdr(ws, row, cols):
    for i, (name, w) in enumerate(cols, 1):
        c = ws.cell(row, i, name)
        c.font, c.fill, c.border, c.alignment = FH, FILL_H, BD, AH
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22


def build_tracker(ws, tasks, statuses):
    ws.merge_cells("A1:K1")
    t = ws.cell(1, 1, "Project Tracker"); t.font = FT; t.alignment = AH
    cols = [("#", 5), ("Task", 30), ("Group", 18), ("Owner", 14), ("Status", 15),
            ("Start", 13), ("End", 13), ("Duration (d)", 12), ("% Complete", 12)]
    _hdr(ws, 3, cols)
    for r, task in enumerate(tasks, start=4):
        vals = [r - 3, task.get("task", ""), task.get("group", ""),
                task.get("owner", ""), task.get("status", ""),
                task["_start"], task["_end"], f"=G{r}-F{r}+1", task.get("percent", 0)]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = FB; c.border = BD
            c.alignment = AL if col in (2, 3) else AC
        ws.cell(r, 6).number_format = ws.cell(r, 7).number_format = DATE_FMT
        ws.row_dimensions[r].height = 30
    last = len(tasks) + 3
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(statuses), allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"E4:E{last}")
    for name, (fill, text, _) in STATUS.items():
        ws.conditional_formatting.add(
            f"E4:E{last}",
            FormulaRule(formula=[f'$E4="{name}"'], stopIfTrue=False,
                        fill=PatternFill("solid", fgColor=fill),
                        font=Font(name="Helvetica", size=10.5, color=text)))
    ws.conditional_formatting.add(
        f"I4:I{last}",
        DataBarRule(start_type="num", start_value=0, end_type="num",
                    end_value=100, color="10B981"))
    ws.freeze_panes = "A4"


def build_gantt(ws, tasks, scale, sheets_native):
    lo = min(t["_start"] for t in tasks)
    hi = max(t["_end"] for t in tasks)
    buckets = _buckets(lo, hi, scale)
    ws.merge_cells("A1:E1")
    t = ws.cell(1, 1, f"Gantt — {scale}ly"); t.font = FT; t.alignment = AH
    lead = [("Task", 28), ("Owner", 13), ("Start", 12), ("End", 12), ("Status", 14)]
    _hdr(ws, 3, lead)
    base = len(lead) + 1  # first bucket column index (F = 6)
    for j, b in enumerate(buckets):
        c = ws.cell(3, base + j, b)
        c.font, c.fill, c.border, c.alignment = FH, FILL_H, BD, AC
        c.number_format = "mmm dd" if scale != "month" else "mmm yyyy"
        ws.column_dimensions[get_column_letter(base + j)].width = 5 if scale == "day" else 9
    first, lastcol = get_column_letter(base), get_column_letter(base + len(buckets) - 1)
    for r, task in enumerate(tasks, start=4):
        row = [task.get("task", ""), task.get("owner", ""), task["_start"],
               task["_end"], task.get("status", "")]
        for col, v in enumerate(row, 1):
            c = ws.cell(r, col, v); c.font = FB; c.border = BD
            c.alignment = AL if col == 1 else AC
        ws.cell(r, 3).number_format = ws.cell(r, 4).number_format = DATE_FMT
        ws.row_dimensions[r].height = 22
    last = len(tasks) + 3
    grid = f"{first}4:{lastcol}{last}"
    end_expr = _bucket_end(first, scale)
    for name, (_, _, bar) in STATUS.items():
        ws.conditional_formatting.add(
            grid,
            FormulaRule(
                formula=[f'AND($E4="{name}",$C4<={end_expr},$D4>={first}$3)'],
                stopIfTrue=True, fill=PatternFill("solid", fgColor=bar)))
    # Today marker on the header row (column fill is robust across Excel + Sheets).
    hdr_end = _bucket_end(first, scale).replace("$3", "3")
    ws.conditional_formatting.add(
        f"{first}3:{lastcol}3",
        FormulaRule(formula=[f"AND({first}3<=TODAY(),TODAY()<={hdr_end})"],
                    stopIfTrue=False, fill=PatternFill("solid", fgColor="1F2937"),
                    font=Font(name="Helvetica", size=10, bold=True, color="FFFFFF")))
    if sheets_native:
        sc = base + len(buckets)
        ws.cell(3, sc, "Bar (Sheets)").font = FH
        ws.column_dimensions[get_column_letter(sc)].width = 22
        span = max((hi - lo).days, 1)
        for r, task in enumerate(tasks, start=4):
            off = (task["_start"] - lo).days
            dur = (task["_end"] - task["_start"]).days + 1
            ws.cell(r, sc, f'=SPARKLINE({{{off},{dur}}},{{"charttype","bar";'
                           f'"max",{span};"color1","white";"color2","#F59E0B"}})')
    ws.freeze_panes = f"{first}4"


def build_dashboard(ws, n, last_row):
    ws.merge_cells("A1:B1")
    t = ws.cell(1, 1, "Dashboard"); t.font = FT; t.alignment = AH
    rows = [
        ("Total tasks", n),
        ("Weighted % complete",
         f"=ROUND(SUMPRODUCT(Tracker!H4:H{last_row},Tracker!I4:I{last_row})/"
         f"SUM(Tracker!H4:H{last_row}),0)"),
        ("Overdue (past End, not Done)",
         f'=SUMPRODUCT((Tracker!G4:G{last_row}<TODAY())*'
         f'(Tracker!E4:E{last_row}<>"Done"))'),
    ]
    for i, (label, val) in enumerate(rows, start=3):
        a = ws.cell(i, 1, label); a.font = FB; a.border = BD; a.alignment = AH
        b = ws.cell(i, 2, val); b.font = FH; b.border = BD; b.alignment = AC
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.cell(len(rows) + 4, 1, "Weighted by task duration. Updates live on edit.").font = FCAP


def main(argv: list[str]) -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("spec", help="path to the JSON spec")
    ap.add_argument("-o", "--out", default="tracker.xlsx")
    ap.add_argument("--timescale", choices=("day", "week", "month"), default=None)
    ap.add_argument("--sheets-native", action="store_true",
                    help="add a SPARKLINE bar column (renders in Sheets, #NAME? in Excel)")
    args = ap.parse_args(argv)

    spec = json.loads(Path(args.spec).read_text(encoding="utf-8"))
    tasks = _norm(spec.get("tasks") or [])
    if not tasks:
        raise SystemExit("spec has no tasks")
    scale = args.timescale or spec.get("timescale", "week")
    statuses = spec.get("statuses") or DEFAULT_STATUSES

    wb = Workbook(); wb.remove(wb.active)
    build_tracker(wb.create_sheet("Tracker"), tasks, statuses)
    build_gantt(wb.create_sheet("Gantt"), tasks, scale, args.sheets_native)
    build_dashboard(wb.create_sheet("Dashboard"), len(tasks), len(tasks) + 3)
    if spec.get("title"):
        wb["Tracker"].cell(1, 1, spec["title"])
    wb.save(args.out)
    print(f"wrote {args.out} — {len(tasks)} tasks, {scale} Gantt")


if __name__ == "__main__":
    main(sys.argv[1:])
